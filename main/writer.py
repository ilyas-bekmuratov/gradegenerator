import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import config
import sys
from copy import copy


def extend_day_columns(sheet, num_copies, is_last_quarter=False, has_exam=False, is_dod=False):
    daily_grade_col_idx = column_index_from_string(config.daily_grade_col)
    max_col_letter = config.dod_hw_col if is_dod else config.hw_col
    max_col = column_index_from_string(max_col_letter)
    styles_widths = {}
    last_real_width = 13.0
    for col_idx in range(daily_grade_col_idx, max_col + 1):
        col_letter = get_column_letter(col_idx)
        style, width = read_styles_and_width(sheet, col_letter)
        if width == 13.0:
            width = last_real_width
        else:
            last_real_width = width
        styles_widths[col_idx] = style, width

    # print(f"   styles_widths uses columns = {list(styles_widths.keys())}")

    # print_widths(sheet, "\ninitial")
    # print(f"merged ranges = {list(sheet.merged_cells.ranges)}")

    yearly_grade_idx = column_index_from_string(config.yearly_grade_col)
    cols_to_delete = []
    if not is_dod:
        if not is_last_quarter:  # delete the final grade, exam, and summary grade columns
            print("      -> not the last quarter removed 3 columns")
            cols_to_delete = [yearly_grade_idx,
                              yearly_grade_idx + 1,
                              yearly_grade_idx + 2]
        elif not has_exam:  # delete the exam and summary grade columns
            print("      -> has no exam, removed 2 columns")
            cols_to_delete = [yearly_grade_idx + 1,
                              yearly_grade_idx + 2]

    new_merges = get_merges_to_restore(cols_to_delete, sheet, num_copies, is_last_quarter, has_exam, is_dod)

    if len(cols_to_delete) > 0:
        sheet.delete_cols(cols_to_delete[0], len(cols_to_delete))
    for col in cols_to_delete:
        del styles_widths[col]
    # print_widths(sheet, f"after deletion of {len(cols_to_delete)} columns at index {yearly_grade_idx}")
    # print(f"   after deletion styles_widths uses columns = {list(styles_widths.keys())}")

    sheet.insert_cols(daily_grade_col_idx, num_copies - 1)
    # print_widths(sheet, f"after insertion of {num_copies - 1} columns")

    for col_idx in range(daily_grade_col_idx, daily_grade_col_idx + num_copies):
        current_col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[current_col_letter].custom_width = True
        styles, sheet.column_dimensions[current_col_letter].width = styles_widths[daily_grade_col_idx]
        for row_idx, style_array in styles.items():
            sheet.cell(row=row_idx, column=col_idx)._style = style_array

    end_index = max_col + num_copies - len(cols_to_delete)
    for col_idx in range(daily_grade_col_idx + num_copies, end_index):
        current_col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[current_col_letter].custom_width = True
        index_to_get_styles = col_idx - num_copies + 1
        if len(cols_to_delete) > 0 and index_to_get_styles >= min(cols_to_delete):
            index_to_get_styles += len(cols_to_delete)
        # print(f"   index_to_get_styles is {index_to_get_styles} is applied to {current_col_letter}")
        styles, sheet.column_dimensions[current_col_letter].width = styles_widths[index_to_get_styles]
        for row_idx, style_array in styles.items():
            sheet.cell(row=row_idx, column=col_idx)._style = style_array

    for merge_str in new_merges:
        sheet.merge_cells(merge_str)
    # print_widths(sheet, "after merging back")
    return sheet


def read_styles_and_width(sheet, col: str):
    styles = {}
    width = sheet.column_dimensions[col].width

    for row_idx in range(1, config.max_row):
        cell = sheet[f'{col}{row_idx}']
        if cell.has_style:
            styles[row_idx] = copy(cell._style)

    return styles, width


def print_widths(sheet, message):
    widths = {}
    for i in range(1, sheet.max_column):
        letter = get_column_letter(i)
        widths[letter] = sheet.column_dimensions[letter].width
    print(f"\n{message}\ncolumn widths = {widths}")


def test(source_file, output_file, num_copies):
    workbook = None
    try:
        workbook = openpyxl.load_workbook(source_file)
    except FileNotFoundError:
        print(f"Error: The template file '{source_file}' was not found.")
        return

    sheetname = "9F - казахский язык и лите - Q4"
    sheet = workbook[sheetname]
    print_widths(sheet, "after saving")
    print(sheet.column_dimensions["AP"].width)
    print(sheet.column_dimensions["AU"].width)

    # extend_day_columns(sheet, num_copies)
    # print(sheet.column_dimensions["AJ"].width)

    # print_widths(sheet)

    # workbook.save(output_file)
    print(f"\nSuccessfully created '{output_file}' with {num_copies} formatted columns.")


def get_merges_to_restore(cols_to_delete, sheet, num_copies, is_last_quarter=False, has_exam=False, is_dod=False):
    daily_grade_col_idx = column_index_from_string(config.daily_grade_col)
    new_merges = []
    dates_idx = 1 if is_dod else column_index_from_string(config.yearly_grade_col)

    for merged_range in list(sheet.merged_cells.ranges):
        if daily_grade_col_idx > merged_range.min_col:
            continue

        sheet.unmerge_cells(str(merged_range))
        if merged_range.min_col in cols_to_delete or merged_range.max_col in cols_to_delete:
            continue
        try:
            offset = -1
            if (len(cols_to_delete) > 0
                    and (merged_range.min_col >= cols_to_delete[0]
                         or merged_range.max_col >= cols_to_delete[-1])):
                offset -= len(cols_to_delete)
            new_min_col = merged_range.min_col + num_copies + offset
            new_max_col = merged_range.max_col + num_copies + offset
            new_range_str = (f"{get_column_letter(new_min_col)}{merged_range.min_row}" +
                             f":{get_column_letter(new_max_col)}{merged_range.max_row}")
            new_merges.append(new_range_str)
        except Exception as e:
            print(f"warning during merge manipulation: {e}")

    return new_merges


def set_column_width_by_string(file_path, search_string, new_width):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: File not found at '{file_path}'")
        return
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    changes_made = False
    print(f"Loading '{file_path}'...")
    print(f"Searching for string: '{search_string}'")

    # Iterate through each worksheet in the workbook
    for ws in wb.worksheets:
        columns_to_change = set()

        # Iterate through all cells in the sheet to find the search string
        for row in ws.iter_rows():
            for cell in row:
                # Check if cell value is a string and contains the search string
                if cell.value is not None and search_string in str(cell.value):
                    # Get the column letter (e.g., 'A', 'B', 'AA')
                    col_letter = get_column_letter(cell.column)
                    columns_to_change.add(col_letter)

        # After checking all cells, apply the width changes for this sheet
        if columns_to_change:
            print(f"  -> Found string in sheet '{ws.title}'.")
            print(f"     Setting width of columns {', '.join(sorted(columns_to_change))} to {new_width}")

            for col_letter in columns_to_change:
                ws.column_dimensions[col_letter].width = new_width

            changes_made = True

    # Save the workbook only if we actually made changes
    if changes_made:
        try:
            wb.save(file_path)
            print(f"\nSuccessfully updated column widths and saved '{file_path}'.")
        except PermissionError:
            print(f"\nError: Could not save '{file_path}'.")
            print("Please make sure the file is not open in Excel.")
        except Exception as e:
            print(f"\nError saving file: {e}")
    else:
        print(f"\nString '{search_string}' was not found in any sheet. No changes made.")


def test2():
    file_path = "reports/testjournals.xlsx"
    search_string = "Темы"
    new_width = 20.37
    if new_width <= 0:
        print("Width must be a positive number.")
        return

    set_column_width_by_string(file_path, search_string, new_width)


if __name__ == "__main__":
    test2()
