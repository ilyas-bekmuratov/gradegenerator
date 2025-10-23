"""
Generates plausible midterm and final exam scores for a given final grade mark
by reverse-engineering the grading process.
### How to Use the Script
0.  **Install libraries:** If you don't have them, open your terminal or command prompt and run:
    `pip install pandas numpy odfpy openpyxl`
1.  **Run the script:** Execute the file from your terminal while inside the folder with these scripts:
     `python main.py`
2. You can change the input by modifying the `config.py` file.
"""

import pandas as pd
import os
import config
import grade_generator as gg
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string
import class_extractor
import topic_extractor
import timetable_extractor
import re
from collections import defaultdict
import random
import helper
import writer
from typing import List, Dict
from classes import Class, Subject


def extract_all_data(class_str: str = "", is_dod=False):
    all_classes_dict = timetable_extractor.extract_class_subjects(class_name=class_str, is_dod=is_dod)
    print(f"class {class_str} has subjects: {all_classes_dict[class_str].subjects.values()}")
    topic_extractor.extract_all_topics_and_hw(all_classes_dict, class_name=class_str, is_dod=is_dod)
    class_extractor.extract_grades_and_classes(all_classes_dict, class_name=class_str)
    return all_classes_dict


def main(is_dod=False):
    all_days_in_year = config.all_days_in_each_quarter
    # all_days_in_year = timetable_extractor.extract_days()
    all_classes_dict = extract_all_data(is_dod=is_dod)

    # --- Group classes by parallel (grade level) ---
    grouped_classes = defaultdict(list)
    for class_name, class_obj in all_classes_dict.items():
        if class_obj is None:
            continue
        match = re.match(r'^\d+', class_name)
        if match:
            parallel = match.group(0)
            grouped_classes[parallel].append(class_obj)

    os.makedirs(config.output_dir, exist_ok=True)

    # --- Loop through each parallel group and create a separate file ---
    for parallel, classes_in_parallel in grouped_classes.items():
        if parallel == "1":
            continue
        output_filename = f"journal {parallel}.xlsx"
        filepath = os.path.join(config.output_dir, output_filename)
        print(f"\n{'='*20} PROCESSING PARALLEL {parallel} {'='*20}")

        workbook = None
        template_path = config.template_path
        try:
            if os.path.exists(filepath):
                workbook = openpyxl.load_workbook(filepath)
                print(f"Successfully loaded existing report from '{filepath}'.")
            else:
                workbook = openpyxl.load_workbook(template_path)
                print(f"Creating new report for parallel {parallel} from template.")

        except FileNotFoundError:
            print(f"Error: Template file not found at '{template_path}'.")
            continue
        except Exception as e:
            print(f"An error occurred while loading the workbook for parallel {parallel}: {e}")
            continue

        for current_class in classes_in_parallel:
            process_class(workbook, current_class, all_days_in_year, is_dod)

        try:
            print("\nCleaning up final workbook...")
            workbook.remove(workbook[config.template_sheet_name])
            workbook.remove(workbook[config.dod_template_sheet_name])

            workbook.save(filepath)
            print(f"\nSuccessfully saved the complete report to '{filepath}'.")
        except Exception as e:
            print(f"\nAn error occurred while saving the file '{filepath}': {e}")


def process_class(
        workbook,
        current_class: Class,
        all_days_in_year: Dict[int, List[str]],
        is_dod=False
):
    for subject_name, subject in current_class.subjects.items():
        print(f"\n--- Processing Subject: {subject_name} ({subject.hours()}h/w) for class {current_class.name} ---")

        class_number_str = re.match(r'^\d+', current_class.name).group(0)
        class_number = int(class_number_str)

        split = 7 if (class_number >= 5 and subject.has_exam) else 5
        split_grades: list[list[int]] = helper.split_string_by_pattern(subject.grades, split)

        for i in range(4):
            quarter_num = i + 1
            print(split_grades[i])
            quarter(workbook, current_class, quarter_num, subject, split_grades, all_days_in_year, is_dod)
            if is_dod:
                break


def quarter(
        workbook,
        current_class: Class,
        quarter_num: int,
        subject: Subject,
        split_grades: list[list[int]],
        all_days_in_each_quarter: Dict[int, List[str]] = config.all_days_in_each_quarter,
        is_dod=False
):
    chrome_length = len(f"{current_class.name} -  - Q{quarter_num}")
    max_subject_len = 31 - chrome_length
    short_subject_name = subject.name[:max_subject_len] if len(subject.name) > max_subject_len else subject.name
    output_sheet_name = f"{current_class.name} - {short_subject_name} - Q{quarter_num}"

    is_art = False
    for art in config.art:
        if art in subject.name:
            is_art = True
            break
    is_boys_art = is_art and config.art_boys[current_class.is_kz] in subject.name
    is_girls_art = is_art and config.art_girls[current_class.is_kz] in subject.name
    if is_boys_art and is_girls_art:
        print(f"Warning art subject {subject} has boys and girls mixed up")

    # Get the original full lists from the class object
    student_list = current_class.students
    gender_list = current_class.genders

    # Prepare new lists to hold the filtered data
    filtered_students = []
    filtered_split_grades = [[] for _ in range(len(split_grades))]

    # print(f"test:   gender list lenth is {len(gender_list)} and student list length is {len(student_list)}")
    if is_art and (is_boys_art or is_girls_art) and len(gender_list) == len(student_list):
        print(f"  -> Applying gender filter for '{subject.name}'")
        for idx, student in enumerate(student_list):
            is_boy = gender_list[idx]

            if (is_boys_art and not is_boy)\
                    or (is_girls_art and is_boy):
                continue

            filtered_students.append(student)

            # We also keep their grades for all quarters/splits
            for q_idx in range(len(split_grades)):
                if idx < len(split_grades[q_idx]):  # Safety check
                    filtered_split_grades[q_idx].append(split_grades[q_idx][idx])
    else:
        # No filter, just use the original lists
        filtered_students = student_list
        filtered_split_grades = split_grades

    # print(f"test:   filtered students: {filtered_students}\n   split grades {split_grades}\n   filtered grades {filtered_split_grades}")

    quarter_grades = filtered_split_grades[quarter_num - 1]

    skip_week = is_dod and subject.name in config.two_per_month
    if is_dod:
        quarter_dates = helper.get_dod_days(subject, all_days_in_each_quarter, skip_week)
    else:
        quarter_dates = helper.get_days_this_quarter(subject, quarter_num, all_days_in_each_quarter)

    total_hours_this_quarter = len(quarter_dates)

    print(f"\n  -> quarter {quarter_num} has grades: {quarter_grades}")
    if total_hours_this_quarter == 0:
        print(f"\n  -> Skipping Quarter {quarter_num} (no lessons).\n")
        return

    print(f"  -> Generating data for Quarter {quarter_num}'...")
    results = []

    num_midterms_for_df = 1 if subject.hours() == 1 else config.num_midterms

    for grade in quarter_grades:
        if grade in [0, 1]:  # Handle blank and pass/fail
            pass_fail_text = ""
            if grade == 1:
                pass_fail_text = "есп" if current_class.is_kz else "зач"

            blank_data = {
                "Input Grade": pass_fail_text,
                "СОр Scores (Midterms)": [''] * num_midterms_for_df,
                "СОч Score (Final)": '', "Adjusted СОр %": '', "Actual СОч %": '',
                "Generated Total %": '', "Penalty/Bonus Applied": 0
            }
            results.append(blank_data)
        elif grade in config.grade_bands:
            generated_data = gg.generate_plausible_grades(grade, subject, quarter_num)
            results.append(generated_data)

    if not results and subject.name not in config.no_grades:
        return
    elif not results and subject.name in config.no_grades:
        blank_data = {
            "Input Grade": '',
            "СОр Scores (Midterms)": [''] * num_midterms_for_df,
            "СОч Score (Final)": '', "Adjusted СОр %": '', "Actual СОч %": '',
            "Generated Total %": '', "Penalty/Bonus Applied": 0
        }
        results.append(blank_data)

    df = pd.DataFrame(results)

    actual_midterm_cols = [f'СОр {j+1}' for j in range(num_midterms_for_df)]
    midterm_df = pd.DataFrame(df['СОр Scores (Midterms)'].tolist(), index=df.index)
    midterm_df.columns = actual_midterm_cols

    template_midterm_cols = [f'СОр {j+1}' for j in range(config.max_midterms)]
    midterm_df = midterm_df.reindex(columns=template_midterm_cols)

    df = pd.concat([midterm_df, df.drop(columns=['СОр Scores (Midterms)'])], axis=1)

    max_sop_weight = config.weights['sop']
    max_so4_weight = config.weights['so4']
    final_df = df.rename(columns={
        'СОч Score (Final)': 'Балл СО за четв.',
        'Adjusted СОр %': f'% СОр (макс. {max_sop_weight}%)',
        'Actual СОч %': f'% СОч (макс. {max_so4_weight}%)',
        'Generated Total %': 'Сумма %', 'Input Grade': 'Оценка за четверть'
    })
    column_order = (
            template_midterm_cols +
            ['Балл СО за четв.', f'% СОр (макс. {max_sop_weight}%)',
             f'% СОч (макс. {max_so4_weight}%)', 'Сумма %', 'Оценка за четверть']
    )
    final_df = final_df[column_order]

    template_sheet_name = config.dod_template_sheet_name if is_dod else config.template_sheet_name

    if output_sheet_name in workbook.sheetnames:
        sheet = workbook[output_sheet_name]
        print(f"  -> Found existing sheet: '{output_sheet_name}'. Overwriting data.")
    else:
        if template_sheet_name not in workbook.sheetnames:
            print(f"  -> ERROR: Template sheet '{template_sheet_name}' not found. Skipping.")
            return
        template_sheet = workbook[template_sheet_name]
        sheet = workbook.copy_worksheet(template_sheet)
        sheet.title = output_sheet_name
        print(f"  -> Created sheet '{output_sheet_name}' "
              f"from template '{template_sheet_name}' for {subject.hours()} hours a week.")

    [student_start_row, student_start_col] = config.student_name_cell
    for idx, student_name in enumerate(filtered_students):
        sheet.cell(row=student_start_row + idx, column=student_start_col, value=student_name)

    [subject_teacher_cell_row, subject_teacher_cell_col] = config.subject_teacher_cell
    title = f"Наименование предмета: {subject.name.capitalize()} Преподаватель: {subject.teacher}"
    sheet.cell(row=subject_teacher_cell_row, column=subject_teacher_cell_col, value=title)

    [quarter_num_cell_row, quarter_num_celll_col] = config.quarter_num_cell
    quarter_text = f"Расчет оценки за {quarter_num}-четверть"
    sheet.cell(row=quarter_num_cell_row, column=quarter_num_celll_col, value=quarter_text)

    rows = dataframe_to_rows(final_df, index=False, header=False)
    col_letter = config.dod_grade_col if is_dod else config.quarter_grade_col
    quarter_grade_start_col = column_index_from_string(col_letter)

    for r_idx, row_data in enumerate(rows, config.start_row):
        for c_idx, value in enumerate(row_data, quarter_grade_start_col):
            sheet.cell(row=r_idx, column=c_idx, value=value if not pd.isna(value) else None)
    print(f"  -> Wrote main grade data for {len(final_df)} students.")

    date_col_letter = config.dod_date_col if is_dod else config.date_col
    dates_start_col = column_index_from_string(date_col_letter)
    topic_col_letter = config.dod_topic_col if is_dod else config.topic_col
    topics_start_col = column_index_from_string(topic_col_letter)
    daily_grades_start_col = column_index_from_string(config.daily_grade_col)

    quarter_topic_start_index = helper.get_quarter_start_index(subject, quarter_num)
    quarter_topic_end_index = min(len(subject.topics)//4, total_hours_this_quarter)
    # --- Topic and Homework Distribution Logic ---
    print(f"  -> Placing {total_hours_this_quarter} dates, topics, homework starting from {quarter_topic_start_index}")
    quarter_topics = subject.topics[quarter_topic_start_index:quarter_topic_start_index+quarter_topic_end_index]
    quarter_hw = subject.homework[quarter_topic_start_index:quarter_topic_start_index+quarter_topic_end_index]

    for idx, date in enumerate(quarter_dates):
        sheet.cell(row=config.start_row + idx, column=dates_start_col, value=date[:5])

    for idx, topic in enumerate(quarter_topics):
        sheet.cell(row=config.start_row + idx, column=topics_start_col, value=topic)

    for idx, hw in enumerate(quarter_hw):
        sheet.cell(row=config.start_row + idx, column=topics_start_col+1, value=hw)

    if quarter_num == 4:
        yearly_grade_col = quarter_grade_start_col + config.quarter_to_dates_offset - 3
        print(f"     -> quarter 4 must have yearly grades")
        for idx, grade in enumerate(filtered_split_grades[4]):
            pass_fail_text = str(grade)
            if grade == 1:
                pass_fail_text = "есп" if current_class.is_kz else "зач"
            sheet.cell(row=config.start_row + idx, column=yearly_grade_col, value=pass_fail_text)
        if subject.has_exam:
            for idx, grade in enumerate(filtered_split_grades[5]):
                sheet.cell(row=config.start_row + idx, column=yearly_grade_col+1, value=grade)
            for idx, grade in enumerate(filtered_split_grades[6]):
                sheet.cell(row=config.start_row + idx, column=yearly_grade_col+2, value=grade)

    # --- Daily Grade Generation Logic ---
    is_last_quarter = quarter_num == 4
    sheet = writer.extend_day_columns(sheet, total_hours_this_quarter, is_last_quarter, subject.has_exam)
    month = ""
    for idx, date in enumerate(quarter_dates):
        sheet.cell(row=config.dates_row, column=daily_grades_start_col + idx, value=date[:2])
        this_month = helper.get_month_from_date(date)
        if this_month != month:
            month = this_month
            sheet.cell(row=config.months_row, column=daily_grades_start_col + idx, value=month)
    print(f"  -> Extended the table by {total_hours_this_quarter} columns")

    if subject.name in config.no_grades:
        print(f"subject {subject.name} has no grades")
        return

    num_grades_to_place = int(total_hours_this_quarter * config.daily_grade_density)
    available_cols = list(range(daily_grades_start_col, quarter_grade_start_col + total_hours_this_quarter - 1))

    for idx, row in df.iterrows():
        bonus = row['Penalty/Bonus Applied']

        quarter_index = quarter_num-1
        if bonus == 0 and subject.hours() == 1:
            if quarter_num == 1 or quarter_num == 3:
                quarter_index += 1  # do not skip for blank or pass/fail grades, use next split grades instead
            else:
                continue

        distribution = config.get_daily_grade_distribution(bonus, filtered_split_grades[quarter_index][idx])
        grades, weights = zip(*distribution.items())
        cols_to_fill = random.sample(available_cols, num_grades_to_place)

        for col in cols_to_fill:
            generated_grade = random.choices(grades, weights=weights, k=1)[0]
            sheet.cell(row=student_start_row + idx, column=col, value=generated_grade)


if __name__ == "__main__":
    main()
